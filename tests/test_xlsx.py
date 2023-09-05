from openpyxl import load_workbook
from conftest import RESOURCE_PATH, xlsx_path

# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def test_xlsx():
    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    cel_rows = sheet.cell(row=3, column=2).value
    print(f"Пересечение строк: {cel_rows}")

    assert cel_rows == 'Mara'